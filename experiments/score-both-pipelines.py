"""Score each pipeline (ask_v1 vs vlm_ask) separately against golden xlsx."""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

OUT_DIR = Path(__file__).parent.parent / "extracted-json"
GOLDEN = OUT_DIR / "model answer.xlsx"

GOLDEN_SHEETS = {
    "2":  "NR Doc2-invoice",
    "6":  "NR Doc6-BL",
    "9":  "NR Doc9-AWB",
    "13": "NR Doc13-Insurancepolicy",
}


def load_golden():
    wb = load_workbook(GOLDEN, data_only=True)
    out = {}
    for sheet_name, _ in GOLDEN_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec, field, value = (
                (row[0] or "").strip(),
                (row[1] or "").strip(),
                (row[2] or "").strip(),
            )
            if not sec:
                continue
            primary = field if field else value
            if not primary:
                continue
            normalized = re.sub(r"\s+", " ", primary).strip(" .,;:")
            rows.append({
                "section": sec,
                "field": field,
                "value": value,
                "search_key": normalized,
                "search_key_lower": normalized.lower(),
            })
        out[sheet_name] = rows
    return out


def score_doc(ask_objs, golden_rows):
    ask_values = []
    for obj in ask_objs:
        for k, v in obj.items():
            if v:
                ask_values.append((k, str(v).strip().lower()))

    def _match(needle):
        if not needle:
            return False, None
        nlow = needle.lower()
        for k, ask_v in ask_values:
            if nlow in ask_v:
                return True, k
        threshold = max(10, int(len(nlow) * 0.7))
        for k, ask_v in ask_values:
            if len(nlow) >= 15 and nlow[:threshold] in ask_v:
                return True, f"{k} (prefix)"
        words = [w for w in re.split(r"[\s,]+", nlow) if len(w) >= 3]
        if len(words) >= 2:
            for k, ask_v in ask_values:
                if all(w in ask_v for w in words):
                    return True, f"{k} (words)"
        return False, None

    hits, misses = [], []
    for row in golden_rows:
        ok, where = _match(row["search_key_lower"])
        if ok:
            hits.append({**row, "found_in": where})
        else:
            misses.append(row)
    total = len(hits) + len(misses)
    return {"total_golden": total, "hits": hits, "misses": misses, "hit_rate": (len(hits) / total) if total else 0.0}


def find_sheet(slug):
    for sheet, pat in GOLDEN_SHEETS.items():
        pat_us = pat.replace(" ", "_")
        slug_clean = re.sub(r"[().pdf]", "", slug)
        if pat_us.lower() in slug.lower() or pat_us.lower() in slug_clean.lower():
            return sheet
    return None


def score_pipeline(prefix, label):
    golden = load_golden()
    json_files = sorted(OUT_DIR.glob(f"{prefix}__*.json"))
    print(f"=== {label} ({prefix}__*.json) ===")
    scored = []
    for jf in json_files:
        with open(jf, encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list) or not data:
            continue
        # Extract slug from filename
        parts = jf.stem.split("__")
        slug = "__".join(parts[2:]) if len(parts) >= 3 else parts[-1]
        sheet = find_sheet(slug)
        if sheet is None:
            continue
        s = score_doc(data, golden[sheet])
        scored.append({"file": jf.name, "slug": slug, "sheet": sheet, "score": s, "data": data})
        print(f"  {slug:55s}  sheet={sheet:>3s}  {s['hit_rate']*100:>5.1f}%  ({len(s['hits'])}/{s['total_golden']})")
    if scored:
        avg = sum(s["score"]["hit_rate"] for s in scored) / len(scored)
        print(f"  AGGREGATE: {avg*100:.1f}% across {len(scored)} golden-mapped docs")
    out_path = OUT_DIR / f"SCORED__{prefix}.json"
    out_path.write_text(json.dumps({
        "label": label,
        "ts": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "aggregate_pct": round(avg * 100, 1) if scored else None,
        "n": len(scored),
        "per_doc": [
            {
                "file": s["file"],
                "filename_slug": s["slug"],
                "sheet": s["sheet"],
                "hit_rate_pct": round(s["score"]["hit_rate"] * 100, 1),
                "hits": [{"section": h["section"], "field": h["field"], "found_in": h["found_in"]} for h in s["score"]["hits"]],
                "misses": [{"section": m["section"], "field": m["field"]} for m in s["score"]["misses"]],
            }
            for s in scored
        ],
    }, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  Wrote {out_path.name}\n")
    return scored


if __name__ == "__main__":
    ask_scored = score_pipeline("ask_v1", "Standard pipeline + Gemma 4 v1 Ask")
    vlm_scored = score_pipeline("vlm_ask", "Qwen3-VL pipeline (markdown) + Gemma 4 v1 Ask")

    # Side-by-side
    print("=" * 80)
    print("SIDE-BY-SIDE COMPARISON")
    print("=" * 80)
    print(f"{'Doc':<35s} {'Standard':>12s} {'VLM-direct':>12s}  {'Delta':>8s}")
    by_sheet_std = {s["sheet"]: s for s in ask_scored}
    by_sheet_vlm = {s["sheet"]: s for s in vlm_scored}
    for sheet in GOLDEN_SHEETS:
        if sheet in by_sheet_std or sheet in by_sheet_vlm:
            s_hr = by_sheet_std.get(sheet, {}).get("score", {}).get("hit_rate", 0) * 100
            v_hr = by_sheet_vlm.get(sheet, {}).get("score", {}).get("hit_rate", 0) * 100
            delta = v_hr - s_hr
            print(f"  {GOLDEN_SHEETS[sheet]:<35s} {s_hr:>11.1f}% {v_hr:>11.1f}%  {delta:>+7.1f}")
    s_avg = sum(s["score"]["hit_rate"] for s in ask_scored) / len(ask_scored) * 100
    v_avg = sum(s["score"]["hit_rate"] for s in vlm_scored) / len(vlm_scored) * 100
    print(f"  {'AGGREGATE':<35s} {s_avg:>11.1f}% {v_avg:>11.1f}%  {v_avg - s_avg:>+7.1f}")
