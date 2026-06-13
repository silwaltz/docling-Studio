"""3-way comparison: Standard / VLM-md-fallback / VLM-json-direct."""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

OUT_DIR = Path(__file__).parent.parent / "extracted-json"
GOLDEN = OUT_DIR / "model answer.xlsx"

GOLDEN_SHEETS = {
    "2":  "NR Doc2-invoice",
    "6":  "NR Doc6-BL",
    "9":  "NR Doc9-AWB",
    "13": "NR Doc13-Insurancepolicy",
}


def load_golden():
    wb = load_workbook(GOLDEN, data_only=True)
    out = {}
    for sheet_name, _ in GOLDEN_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec, field, value = (
                (row[0] or "").strip(),
                (row[1] or "").strip(),
                (row[2] or "").strip(),
            )
            if not sec:
                continue
            primary = field if field else value
            if not primary:
                continue
            normalized = re.sub(r"\s+", " ", primary).strip(" .,;:")
            rows.append({
                "section": sec,
                "field": field,
                "value": value,
                "search_key": normalized,
                "search_key_lower": normalized.lower(),
            })
        out[sheet_name] = rows
    return out


def score_doc(ask_objs, golden_rows):
    ask_values = []
    for obj in ask_objs:
        for k, v in obj.items():
            if v:
                ask_values.append((k, str(v).strip().lower()))

    def _match(needle):
        if not needle:
            return False, None
        nlow = needle.lower()
        for k, ask_v in ask_values:
            if nlow in ask_v:
                return True, k
        threshold = max(10, int(len(nlow) * 0.7))
        for k, ask_v in ask_values:
            if len(nlow) >= 15 and nlow[:threshold] in ask_v:
                return True, f"{k} (prefix)"
        words = [w for w in re.split(r"[\s,]+", nlow) if len(w) >= 3]
        if len(words) >= 2:
            for k, ask_v in ask_values:
                if all(w in ask_v for w in words):
                    return True, f"{k} (words)"
        return False, None

    hits, misses = [], []
    for row in golden_rows:
        ok, where = _match(row["search_key_lower"])
        if ok:
            hits.append({**row, "found_in": where})
        else:
            misses.append(row)
    total = len(hits) + len(misses)
    return {"total_golden": total, "hits": hits, "misses": misses, "hit_rate": (len(hits) / total) if total else 0.0}


def find_sheet(slug):
    for sheet, pat in GOLDEN_SHEETS.items():
        pat_us = pat.replace(" ", "_")
        slug_clean = re.sub(r"[().pdf]", "", slug)
        if pat_us.lower() in slug.lower() or pat_us.lower() in slug_clean.lower():
            return sheet
    return None


def score_all(prefix, label):
    """Score a pipeline's outputs against golden."""
    golden = load_golden()
    json_files = sorted(OUT_DIR.glob(f"{prefix}__*.json"))
    json_files = [f for f in json_files if "RUN_SUMMARY" not in f.name and "SCORED" not in f.name and not f.name.endswith(".raw.json") and not f.name.endswith(".ERROR.json") and not f.name.endswith(".NOPARSE.json")]
    scored = []
    for jf in json_files:
        with open(jf, encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list) or not data:
            continue
        parts = jf.stem.split("__")
        slug = "__".join(parts[2:]) if len(parts) >= 3 else parts[-1]
        sheet = find_sheet(slug)
        if sheet is None:
            continue
        s = score_doc(data, golden[sheet])
        scored.append({"slug": slug, "sheet": sheet, "score": s})
    avg = sum(s["score"]["hit_rate"] for s in scored) / len(scored) if scored else 0
    return scored, avg


if __name__ == "__main__":
    print("Scoring 3 pipelines...")
    std_scored, std_avg = score_all("ask_v1", "Standard + Gemma 4 v1 Ask")
    vlm_md_scored, vlm_md_avg = score_all("vlm_ask", "VLM-direct (was actually OCR fallback) + Ask")
    vlm_json_scored, vlm_json_avg = score_all("vlm_json", "VLM-direct JSON (qwen3-vl-instruct, NO Ask)")

    # Side-by-side
    print("\n" + "=" * 80)
    print("3-WAY PIPELINE COMPARISON")
    print("=" * 80)
    print(f"{'Doc':<35s} {'Standard+Ask':>14s} {'VLM-fallback+Ask':>18s} {'VLM-json direct':>16s}")
    by_sheet_std = {s["sheet"]: s for s in std_scored}
    by_sheet_md = {s["sheet"]: s for s in vlm_md_scored}
    by_sheet_j = {s["sheet"]: s for s in vlm_json_scored}
    for sheet in GOLDEN_SHEETS:
        s_hr = by_sheet_std.get(sheet, {}).get("score", {}).get("hit_rate", 0) * 100
        m_hr = by_sheet_md.get(sheet, {}).get("score", {}).get("hit_rate", 0) * 100
        j_hr = by_sheet_j.get(sheet, {}).get("score", {}).get("hit_rate", 0) * 100
        print(f"  {GOLDEN_SHEETS[sheet]:<35s} {s_hr:>13.1f}% {m_hr:>17.1f}% {j_hr:>15.1f}%")
    print(f"  {'AGGREGATE':<35s} {std_avg*100:>13.1f}% {vlm_md_avg*100:>17.1f}% {vlm_json_avg*100:>15.1f}%")
    print()
    print("Notes:")
    print("  - Standard+Ask: docling standard pipeline -> Gemma 4 v1 prompt Ask")
    print("  - VLM-fallback+Ask: my previous 'VLM-direct markdown' test — actually fell back to standard+OCR due to wrong model name")
    print("  - VLM-json direct: qwen3-vl:8b-instruct, json mode, output is final (no Ask step)")
    print("  - VLM-json uses the v1 prompt synced to chat.py v1 (after the 2026-06-13 Doc13 fragmentation fix)")
