"""
Ask-pipeline prompt iteration harness.

Calls Ollama directly with a configurable system prompt + a per-doc user message.
Records all outputs to disk and scores them against the golden xlsx.

Why direct Ollama (not /api/documents/:id/chat):
- We want to swap the system prompt without restarting the backend
- We want to swap the model without restarting anything
- chat.py's only logic is: lookup latest analysis, slice to 96k chars, call Ollama
  with our system prompt. We replicate that here and gain iteration speed.

Output layout:
  experiments/prompt-runs/<run-name>/
    config.json             (model + prompt + settings)
    <doc_id>__<slug>.raw    (raw ollama response text)
    <doc_id>__<slug>.json   (parsed JSON, what the frontend would download)
    scores.json             (per-doc + aggregate score)
    summary.md              (human-readable summary)
"""
import argparse
import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path
from typing import Optional

# Force utf-8 on Windows console to avoid cp950/cp1252 crashes on model output
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ---- Config ----
OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
DEFAULT_MODEL = "gemma4:e4b-it-qat"
DOCS_FILE = Path(__file__).parent / "docs_full.json"
GOLDEN_XLSX = Path(__file__).parent.parent / "extracted-json" / "model answer.xlsx"
RUNS_DIR = Path(__file__).parent / "prompt-runs"

# Sheets -> (golden key, doc filename pattern)
GOLDEN_SHEETS = {
    "2":  "NR Doc2-invoice",
    "6":  "NR Doc6-BL",
    "9":  "NR Doc9-AWB",
    "13": "NR Doc13-Insurancepolicy",
}


# ---- Ollama call (non-streaming) ----
def call_ollama(model: str, system: str, user: str, timeout: int = 300) -> tuple[str, dict]:
    """Return (full_response_text, metadata)."""
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        "stream": False,
        "options": {
            "num_ctx": 96_000,
            "num_predict": 96_000,
            "temperature": 0.0,  # deterministic for iteration
        },
    }
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        f"{OLLAMA_HOST}/api/chat",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    t0 = time.time()
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read())
    elapsed = time.time() - t0
    text = data.get("message", {}).get("content", "")
    meta = {
        "model": data.get("model", model),
        "total_duration": data.get("total_duration", 0),
        "load_duration": data.get("load_duration", 0),
        "prompt_eval_count": data.get("prompt_eval_count", 0),
        "eval_count": data.get("eval_count", 0),
        "elapsed_sec": round(elapsed, 2),
    }
    return text, meta


# ---- JSON extraction (mirrors frontend DocAskTab.vue extractJson) ----
def extract_json(text: str) -> Optional[str]:
    """Pull the first parseable JSON object/array from a model response."""
    # 1) code fence
    fenced = re.search(r"```(?:json)?\s*([\s\S]+?)```", text)
    if fenced:
        candidate = fenced.group(1).strip()
        try:
            json.loads(candidate)
            return candidate
        except Exception:
            pass
    # 2) bare JSON
    for open_char, close_char in [("{", "}"), ("[", "]")]:
        start = text.find(open_char)
        if start < 0:
            continue
        depth = 0
        in_str = False
        escape = False
        for i in range(start, len(text)):
            ch = text[i]
            if escape:
                escape = False
                continue
            if in_str:
                if ch == "\\":
                    escape = True
                elif ch == '"':
                    in_str = False
                continue
            if ch == '"':
                in_str = True
            elif ch == open_char:
                depth += 1
            elif ch == close_char:
                depth -= 1
                if depth == 0:
                    candidate = text[start : i + 1]
                    try:
                        json.loads(candidate)
                        return candidate
                    except Exception:
                        break
    return None


def parse_json_array(raw: str) -> list[dict]:
    """Parse model output into a list of dicts. Handles single-object and array output."""
    cleaned = raw.replace("\\_", " ")
    parsed = json.loads(cleaned)
    if isinstance(parsed, dict):
        return [parsed]
    if isinstance(parsed, list):
        return parsed
    return []


# ---- Scoring against golden ----
def load_golden() -> dict[str, list[dict]]:
    """sheet_name -> list of {section, field, value, search_key} rows.

    search_key is the substring we look for in the model output. By default
    it's the Field column (the entity name); for goods where field is the
    description, we use the field too. We also add a normalized version with
    trailing/extra punctuation stripped to make matching more forgiving.
    """
    from openpyxl import load_workbook
    wb = load_workbook(GOLDEN_XLSX, data_only=True)
    out = {}
    for sheet_name, _ in GOLDEN_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec, field, value = (
                (row[0] or "").strip(),
                (row[1] or "").strip(),
                (row[2] or "").strip(),
            )
            if not sec:
                continue
            # The Field column is the entity we want to find in the model output.
            # The Value column is a role/label (e.g. "Shipper", "Port of Loading"),
            # which is metadata about the field, not a separate entity to extract.
            primary = field if field else value
            if not primary:
                continue
            # Normalize: collapse spaces, strip trailing punctuation
            normalized = re.sub(r"\s+", " ", primary).strip(" .,;:")
            rows.append({
                "section": sec,
                "field": field,
                "value": value,
                "search_key": normalized,
                "search_key_lower": normalized.lower(),
            })
        out[sheet_name] = rows
    return out


def score_doc(ask_objs: list[dict], golden_rows: list[dict]) -> dict:
    """For each golden row, check if its `search_key` appears (case-insensitive)
    in any of the model values (the value side of the JSON objects).
    Also tries a 70%-length prefix match for entity names that may have
    punctuation stripped in the model output.
    """
    ask_values = []
    for obj in ask_objs:
        for k, v in obj.items():
            if v:
                ask_values.append((k, str(v).strip().lower()))

    def _match(needle: str) -> tuple[bool, str | None]:
        if not needle:
            return False, None
        nlow = needle.lower()
        for k, ask_v in ask_values:
            if nlow in ask_v:
                return True, k
        # prefix match: 70% of needle chars
        threshold = max(10, int(len(nlow) * 0.7))
        for k, ask_v in ask_values:
            if len(nlow) >= 15 and nlow[:threshold] in ask_v:
                return True, f"{k} (prefix)"
        # word-level match: all significant words of needle present in ask_v
        words = [w for w in re.split(r"[\s,]+", nlow) if len(w) >= 3]
        if len(words) >= 2:
            for k, ask_v in ask_values:
                if all(w in ask_v for w in words):
                    return True, f"{k} (words)"
        return False, None

    hits = []
    misses = []
    for row in golden_rows:
        key = row["search_key_lower"]
        ok, found_in = _match(key)
        if ok:
            hits.append({**row, "found_in": found_in})
        else:
            misses.append(row)
    total = len(hits) + len(misses)
    return {
        "total_golden": total,
        "hits": hits,
        "misses": misses,
        "hit_rate": (len(hits) / total) if total else 0.0,
    }


# ---- Run ----
def run_iteration(name: str, model: str, system_prompt_template: str, user_message: str, max_ctx_chars: int = 96_000):
    RUNS_DIR.mkdir(parents=True, exist_ok=True)
    run_dir = RUNS_DIR / name
    run_dir.mkdir(parents=True, exist_ok=True)
    with open(DOCS_FILE) as f:
        docs = json.load(f)
    golden = load_golden()
    print(f"[{name}] model={model}  docs={len(docs)}  golden_sheets={list(golden)}")

    # Persist config
    (run_dir / "config.json").write_text(json.dumps({
        "name": name,
        "model": model,
        "system_prompt": system_prompt_template,
        "user_message": user_message,
        "max_ctx_chars": max_ctx_chars,
        "ts": datetime.now().isoformat(timespec="seconds"),
    }, indent=2), encoding="utf-8")

    results = []
    for d in docs:
        md = d["md"]
        if len(md) > max_ctx_chars:
            md = md[:max_ctx_chars] + "\n\n[Document truncated for context window]"
        sys_prompt = system_prompt_template.format(context=md)
        print(f"  - {d['filename']:55s} md={d['md_chars']:>6}c  ...", end=" ", flush=True)
        try:
            text, meta = call_ollama(model, sys_prompt, user_message)
        except Exception as e:
            print(f"ERROR: {e}")
            results.append({"doc": d, "error": str(e)})
            continue

        slug = re.sub(r"[^\w.-]+", "_", d["filename"])[:80]
        (run_dir / f"{d['doc_id'][:8]}__{slug}.raw").write_text(text, encoding="utf-8")
        parsed = None
        score = None
        try:
            json_str = extract_json(text)
            if json_str:
                objs = parse_json_array(json_str)
                (run_dir / f"{d['doc_id'][:8]}__{slug}.json").write_text(
                    json.dumps(objs, indent=2, ensure_ascii=False),
                    encoding="utf-8",
                )
                # Find which golden sheet this doc matches
                for sheet, fn_pattern in GOLDEN_SHEETS.items():
                    if fn_pattern in d["filename"]:
                        score = score_doc(objs, golden[sheet])
                        break
                parsed = {"n_objects": len(objs), "score": score}
        except Exception as e:
            parsed = {"error": str(e)}
        print(f"tokens={meta['eval_count']:>5}  {meta['elapsed_sec']}s  {parsed}")
        results.append({"doc": d, "meta": meta, "parsed": parsed, "score": score})

    # Aggregate
    scored = [r for r in results if r.get("score")]
    if scored:
        avg = sum(r["score"]["hit_rate"] for r in scored) / len(scored)
        print(f"\n[{name}] AGGREGATE: {len(scored)} golden-mapped docs, avg hit_rate = {avg*100:.1f}%")
    else:
        avg = None
        print(f"\n[{name}] no golden-mapped docs in this run")

    (run_dir / "scores.json").write_text(json.dumps({
        "name": name,
        "model": model,
        "aggregate_avg_hit_rate": avg,
        "per_doc": [
            {
                "doc_id": r["doc"]["doc_id"],
                "filename": r["doc"]["filename"],
                "score": r["score"],
                "meta": r.get("meta"),
            }
            for r in results if r.get("score") is not None
        ],
        "all_results": [
            {
                "doc_id": r["doc"]["doc_id"],
                "filename": r["doc"]["filename"],
                "has_json": (r.get("parsed") or {}).get("n_objects") is not None,
                "error": r.get("error"),
                "parsed_error": (r.get("parsed") or {}).get("error"),
            }
            for r in results
        ],
    }, indent=2, default=str), encoding="utf-8")

    # Human summary
    lines = [f"# {name}", "", f"**Model:** `{model}`", f"**Avg hit rate:** {avg*100:.1f}%" if avg is not None else "_no golden-mapped docs_", ""]
    for r in results:
        if r.get("score"):
            s = r["score"]
            lines.append(f"## {r['doc']['filename']}")
            lines.append(f"- hit_rate: **{s['hit_rate']*100:.0f}%** ({len(s['hits'])}/{s['total_golden']})")
            if s["misses"]:
                lines.append(f"- misses:")
                for miss in s["misses"]:
                    lines.append(f"  - `{miss['section']}` / `{miss['field'][:60]}`")
            lines.append("")
    (run_dir / "summary.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"\nWrote: {run_dir}")
    return results


# ---- Baseline prompt (current production) ----
BASELINE_PROMPT = (
    """
    Extract every distinct entity from this page and return a single JSON object. "
    "Use exactly these four key prefixes with a numeric suffix starting at "
    "1 (Company Name1, Address1, Shipping Information1, Goods Description1, etc.): "
    "\"Company Name<n>\" = a company, bank, agency, or organization name; "
    "\"Address<n>\" = a postal or physical address (one line per key, multi-line "
    "addresses get multiple Address<n>); "
    "\"Shipping Information<n>\" = port, vessel, container, B/L, AWB, or routing info; "
    "\"Goods Description<n>\" = a description of goods, items, or cargo. "
    "Rules: Output ONLY the JSON object - no markdown, no code fences, no commentary. "
    "Every value MUST be a plain string (no arrays, no nested objects). "
    "Include every distinct entity you find; do not deduplicate. "
    "If a section has no entries on this page, omit its keys entirely. "
    "Replace newlines inside values with a single space. "
    "Document context: {context}"
    """
)

USER_MESSAGE = "Extract all companies, addresses, shipping information, and goods descriptions from the document above as JSON."


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--name", required=True, help="Run name, e.g. baseline / v1 / v2-qwen")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--prompt", default=None, help="Path to a prompt file. Default = baseline.")
    parser.add_argument("--user-message", default=USER_MESSAGE)
    args = parser.parse_args()
    if args.prompt:
        template = Path(args.prompt).read_text(encoding="utf-8")
    else:
        template = BASELINE_PROMPT
    run_iteration(args.name, args.model, template, args.user_message)
