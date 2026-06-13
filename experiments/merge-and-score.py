"""Merge standard+Ask and VLM-json outputs for all 4 golden-mapped docs, then score."""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

OUT_DIR = Path(__file__).parent.parent / "extracted-json"
GOLDEN = OUT_DIR / "model answer.xlsx"

GOLDEN_SHEETS = {
    "2":  "NR Doc2-invoice",
    "6":  "NR Doc6-BL",
    "9":  "NR Doc9-AWB",
    "13": "NR Doc13-Insurancepolicy",
}

# Map (sheet) -> (standard filename slug, vlm filename slug)
DOCS = {
    "2":  ("NR_Doc2-invoice.pdf", "NR_Doc2-invoice.pdf"),
    "6":  ("NR_Doc6-BL.pdf", "NR_Doc6-BL.pdf"),
    "9":  ("NR_Doc9-AWB.pdf", "NR_Doc9-AWB.pdf"),
    "13": ("NR_Doc13-Insurancepolicy.pdf", "NR_Doc13-Insurancepolicy.pdf"),
}


# ---- Merge logic (same as merge-pipelines.py but inline) ----
SECTION_PREFIXES = ("Company Name", "Address", "Shipping Information", "Goods Description")

def get_section(key: str):
    for p in SECTION_PREFIXES:
        if key.startswith(p):
            return p
    return None


def normalize_value(v: str) -> str:
    return re.sub(r"[\s.,;:]+", " ", (v or "").strip().lower())


def is_substring_of(a_norm: str, b_norm: str) -> bool:
    return a_norm == b_norm or (a_norm and a_norm in b_norm)


def dedup_keep_longer(values):
    kept = []  # list of (raw, normalized)
    for v in values:
        if not v or not v.strip():
            continue
        nv = normalize_value(v)
        replace_idx = -1
        skip = False
        for i, (_, kept_n) in enumerate(kept):
            if is_substring_of(nv, kept_n):
                # current value is substring of an existing one; skip current
                skip = True
                break
            elif is_substring_of(kept_n, nv):
                # existing is substring of current; replace existing
                replace_idx = i
                break
        if skip:
            continue
        if replace_idx >= 0:
            kept[replace_idx] = (v.strip(), nv)
        else:
            kept.append((v.strip(), nv))
    return [raw for raw, _ in kept]


def merge_outputs(standard_objs, vlm_objs):
    by_section = {p: [] for p in SECTION_PREFIXES}
    for obj_list in (standard_objs, vlm_objs):
        for obj in obj_list:
            for k, v in obj.items():
                sec = get_section(k)
                if sec is None or v is None:
                    continue
                vs = str(v).strip()
                if vs:
                    by_section[sec].append(vs)
    merged = {}
    for sec, vals in by_section.items():
        deduped = dedup_keep_longer(vals)
        for i, v in enumerate(deduped, start=1):
            merged[f"{sec}{i}"] = v
    return [merged]


# ---- Load golden + scorer ----
def load_golden():
    wb = load_workbook(GOLDEN, data_only=True)
    out = {}
    for sheet_name, _ in GOLDEN_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec, field, value = (
                (row[0] or "").strip(),
                (row[1] or "").strip(),
                (row[2] or "").strip(),
            )
            if not sec:
                continue
            primary = field if field else value
            if not primary:
                continue
            normalized = re.sub(r"\s+", " ", primary).strip(" .,;:")
            rows.append({
                "section": sec,
                "field": field,
                "value": value,
                "search_key": normalized,
                "search_key_lower": normalized.lower(),
            })
        out[sheet_name] = rows
    return out


def score_doc(ask_objs, golden_rows):
    ask_values = []
    for obj in ask_objs:
        for k, v in obj.items():
            if v:
                ask_values.append((k, str(v).strip().lower()))

    def _match(needle):
        if not needle:
            return False, None
        nlow = needle.lower()
        for k, ask_v in ask_values:
            if nlow in ask_v:
                return True, k
        threshold = max(10, int(len(nlow) * 0.7))
        for k, ask_v in ask_values:
            if len(nlow) >= 15 and nlow[:threshold] in ask_v:
                return True, f"{k} (prefix)"
        words = [w for w in re.split(r"[\s,]+", nlow) if len(w) >= 3]
        if len(words) >= 2:
            for k, ask_v in ask_values:
                if all(w in ask_v for w in words):
                    return True, f"{k} (words)"
        return False, None

    hits, misses = [], []
    for row in golden_rows:
        ok, where = _match(row["search_key_lower"])
        if ok:
            hits.append({**row, "found_in": where})
        else:
            misses.append(row)
    total = len(hits) + len(misses)
    return {"total_golden": total, "hits": hits, "misses": misses, "hit_rate": (len(hits) / total) if total else 0.0}


if __name__ == "__main__":
    golden = load_golden()
    print("=" * 90)
    print("MERGED PIPELINE: standard+Ask (union) vlm-json")
    print("=" * 90)
    print()
    scored = []
    for sheet, (std_slug, vlm_slug) in DOCS.items():
        # Find standard+Ask JSON
        std_files = list(OUT_DIR.glob(f"ask_v1__*__{std_slug}.json"))
        vlm_files = list(OUT_DIR.glob(f"vlm_json__*__{vlm_slug}.json"))
        if not std_files or not vlm_files:
            print(f"Sheet {sheet}: missing files std={len(std_files)} vlm={len(vlm_files)}")
            continue
        std_data = json.loads(std_files[0].read_text(encoding="utf-8"))
        vlm_data = json.loads(vlm_files[0].read_text(encoding="utf-8"))
        merged = merge_outputs(std_data, vlm_data)
        s = score_doc(merged, golden[sheet])
        # Save merged output
        out_path = OUT_DIR / f"merged__{std_files[0].name.split('__')[1]}__{std_slug.replace('.pdf','')}.json"
        out_path.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
        scored.append({"sheet": sheet, "slug": std_slug, "score": s, "merged_path": out_path.name})
        print(f"  Sheet {sheet} ({std_slug})")
        print(f"    merged: {sum(len(o) for o in merged)} keys")
        print(f"    {s['hit_rate']*100:>5.1f}% ({len(s['hits'])}/{s['total_golden']})")
    if scored:
        avg = sum(s["score"]["hit_rate"] for s in scored) / len(scored)
        print(f"\n  AGGREGATE: {avg*100:.1f}% across {len(scored)} golden-mapped docs")
    out = {
        "label": "Merged: standard+Ask (union) vlm-json",
        "ts": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "aggregate_pct": round(avg * 100, 1) if scored else None,
        "n": len(scored),
        "per_doc": [
            {
                "sheet": s["sheet"],
                "filename_slug": s["slug"],
                "merged_file": s["merged_path"],
                "hit_rate_pct": round(s["score"]["hit_rate"] * 100, 1),
                "hits": [{"section": h["section"], "field": h["field"], "found_in": h["found_in"]} for h in s["score"]["hits"]],
                "misses": [{"section": m["section"], "field": m["field"]} for m in s["score"]["misses"]],
            }
            for s in scored
        ],
    }
    (OUT_DIR / "SCORED__merged.json").write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n  Wrote SCORED__merged.json")
