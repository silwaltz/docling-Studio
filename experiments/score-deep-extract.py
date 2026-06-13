"""Score the deep-extract (extractMode=deep) outputs against the golden xlsx.

The deep-extract runner saves each analysis as a single JSON file
containing the response payload, where `result.contentJson` is the
*merged* dict (union of standard+Ask and VLM-json). The scorer walks
each golden sheet, takes its `field`/`value` rows, and matches them
against any value in the deep-extract dict.
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

OUT_DIR = Path(__file__).parent.parent / "extracted-json"
GOLDEN = OUT_DIR / "model answer.xlsx"

GOLDEN_SHEETS = {
    "2":  "Doc2",
    "6":  "Doc6",
    "9":  "Doc9",
    "13": "Doc13",
}


def load_golden():
    wb = load_workbook(GOLDEN, data_only=True)
    out = {}
    for sheet_name, _ in GOLDEN_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec, field, value = (
                (row[0] or "").strip(),
                (row[1] or "").strip(),
                (row[2] or "").strip(),
            )
            if not sec:
                continue
            primary = field if field else value
            if not primary:
                continue
            normalized = re.sub(r"\s+", " ", primary).strip(" .,;:")
            rows.append({
                "section": sec,
                "field": field,
                "value": value,
                "search_key": normalized,
                "search_key_lower": normalized.lower(),
            })
        out[sheet_name] = rows
    return out


def find_sheet_from_slug(slug: str) -> str | None:
    for sheet, pat in GOLDEN_SHEETS.items():
        if pat in slug:
            return sheet
    return None


def score_dict(de_dict: dict, golden_rows: list) -> dict:
    """Score a single dict (deep-extract merged output) against golden rows.

    Strategy mirrors the v1 scorer: any value in the dict counts as a
    hit if the golden's search_key (lowercased) appears as a substring,
    or as a 70% prefix for long needles, or as all-significant-words
    for multi-word needles.
    """
    values = []
    for k, v in de_dict.items():
        if v is None:
            continue
        # Values can be strings, numbers, or dicts (after _sanitize_ask_values
        # they should all be strings). Coerce defensively.
        if isinstance(v, dict):
            for sub_k, sub_v in v.items():
                if sub_v:
                    values.append((f"{k}.{sub_k}", str(sub_v).strip().lower()))
        else:
            values.append((k, str(v).strip().lower()))

    def _match(needle: str) -> tuple[bool, str | None]:
        if not needle:
            return False, None
        nlow = needle.lower()
        for k, v in values:
            if nlow in v:
                return True, k
        threshold = max(10, int(len(nlow) * 0.7))
        for k, v in values:
            if len(nlow) >= 15 and nlow[:threshold] in v:
                return True, f"{k} (prefix)"
        words = [w for w in re.split(r"[\s,]+", nlow) if len(w) >= 3]
        if len(words) >= 2:
            for k, v in values:
                if all(w in v for w in words):
                    return True, f"{k} (words)"
        return False, None

    hits, misses = [], []
    for row in golden_rows:
        ok, where = _match(row["search_key_lower"])
        if ok:
            hits.append({**row, "found_in": where})
        else:
            misses.append(row)
    total = len(hits) + len(misses)
    return {
        "total_golden": total,
        "hits": len(hits),
        "misses": len(misses),
        "hit_rate": (len(hits) / total) if total else 0.0,
        "miss_examples": [m["search_key"] for m in misses[:5]],
    }


def main():
    golden = load_golden()
    files = sorted(OUT_DIR.glob("deep_extract__*.json"))
    # Skip the summary file
    files = [f for f in files if "FULL_SUMMARY" not in f.name]
    if not files:
        print("No deep_extract__*.json files found. Run run-full-deep-test.py first.")
        return
    print(f"Scoring {len(files)} deep-extract outputs against {len(GOLDEN_SHEETS)} golden sheets")
    print("=" * 80)
    scored = []
    for jf in files:
        slug = jf.stem.replace("deep_extract__", "")
        sheet = find_sheet_from_slug(slug)
        with open(jf, encoding="utf-8") as f:
            data = json.load(f)
        cj = data.get("result", {}).get("contentJson")
        if not cj:
            print(f"  {slug:50s} no contentJson (status={data.get('result', {}).get('status')})")
            continue
        try:
            de = json.loads(cj) if isinstance(cj, str) else cj
        except Exception as e:
            print(f"  {slug:50s} contentJson parse failed: {e}")
            continue
        if sheet is None:
            # Doc not in golden — still report keys count
            print(f"  {slug:50s} {len(de)} keys (not in golden)")
            continue
        s = score_dict(de, golden[sheet])
        scored.append({"slug": slug, "sheet": sheet, "score": s, "key_count": len(de)})

    print()
    print(f"{'Sheet':<6s} {'Doc':<35s} {'Score':>8s}  {'Hits':>5s}/{'Total':<5s}  {'Keys':>5s}")
    print("-" * 80)
    by_sheet = {s["sheet"]: s for s in scored}
    for sheet, _ in GOLDEN_SHEETS.items():
        s = by_sheet.get(sheet, {}).get("score")
        slug = by_sheet.get(sheet, {}).get("slug", "(not scored)")
        if s:
            print(
                f"  {sheet:<4s} {slug:<35s} {s['hit_rate']*100:>7.1f}%  "
                f"{s['hits']:>5d}/{s['total_golden']:<5d}  "
                f"{by_sheet.get(sheet, {}).get('key_count', 0):>5d}"
            )
        else:
            print(f"  {sheet:<4s} {slug:<35s}    --  no output")
    if scored:
        avg = sum(s["score"]["hit_rate"] for s in scored) / len(scored)
        print("-" * 80)
        print(f"  AGGREGATE  (across {len(scored)} golden-mapped docs)  {avg*100:>6.1f}%")
    print()
    # Per-doc miss examples
    for s in scored:
        ex = s["score"].get("miss_examples", [])
        if ex:
            print(f"  {s['slug']} misses (examples): {ex[:3]}")


if __name__ == "__main__":
    main()
